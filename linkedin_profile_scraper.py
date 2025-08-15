#!/usr/bin/env python3
"""
LinkedIn Profile Scraper with Skill Matching and Excel Export
Extracts profiles, skills, and contact info, then calculates match scores
"""

import logging
import time
import os
import re
from typing import List, Dict, Any, Optional
from datetime import datetime
import pandas as pd
from dataclasses import dataclass, asdict

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from bs4 import BeautifulSoup

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@dataclass
class ProfileData:
    """Store LinkedIn profile data"""
    name: str
    headline: str = ""
    location: str = ""
    profile_url: str = ""
    email: str = ""
    current_company: str = ""
    experience: str = ""
    skills: List[str] = None
    skill_match_score: float = 0.0
    required_skills_matched: List[str] = None
    total_skills_count: int = 0
    about: str = ""
    education: str = ""
    connections: str = ""
    
    def to_dict(self):
        return {
            'Name': self.name,
            'Email': self.email if self.email else 'Not Available',
            'Headline': self.headline,
            'Current Company': self.current_company,
            'Location': self.location,
            'Experience Summary': self.experience,
            'Education': self.education,
            'Skills Matched': ', '.join(self.required_skills_matched) if self.required_skills_matched else '',
            'Match Score (%)': round(self.skill_match_score, 1),
            'Total Skills': self.total_skills_count,
            'All Skills': ', '.join(self.skills) if self.skills else '',
            'About': self.about[:200] + '...' if self.about and len(self.about) > 200 else self.about,
            'Profile URL': self.profile_url,
            'Connections': self.connections
        }


class LinkedInProfileScraper:
    """Scrape LinkedIn profiles with skill matching"""
    
    def __init__(self, email: str = None, password: str = None, headless: bool = False):
        self.email = email or os.getenv('LINKEDIN_EMAIL')
        self.password = password or os.getenv('LINKEDIN_PASSWORD')
        self.headless = headless
        self.driver = None
        self.wait_time = 3
        
        # Target skills for Oracle PL/SQL Developer
        self.target_skills = [
            'Oracle', 'PL/SQL', 'SQL', 'Database', 'Oracle Database',
            'Stored Procedures', 'Triggers', 'Functions', 'Packages',
            'Performance Tuning', 'Query Optimization', 'Data Modeling',
            'ETL', 'Data Warehouse', 'Oracle Forms', 'Oracle Reports',
            'APEX', 'SQL*Plus', 'TOAD', 'SQL Developer', 'Unix',
            'Shell Scripting', 'Python', 'Java', 'JavaScript'
        ]
    
    def setup_driver(self):
        """Setup Chrome driver with anti-detection measures"""
        options = Options()
        
        if self.headless:
            options.add_argument("--headless")
            
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        options.add_argument("--window-size=1920,1080")
        options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36")
        
        self.driver = webdriver.Chrome(options=options)
        self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    
    def login(self) -> bool:
        """Login to LinkedIn"""
        try:
            logger.info("Logging into LinkedIn...")
            self.driver.get("https://www.linkedin.com/login")
            time.sleep(2)
            
            # Enter credentials
            email_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.ID, "username"))
            )
            email_field.send_keys(self.email)
            
            password_field = self.driver.find_element(By.ID, "password")
            password_field.send_keys(self.password)
            
            # Submit
            submit_button = self.driver.find_element(By.XPATH, "//button[@type='submit']")
            submit_button.click()
            
            time.sleep(5)
            
            # Check if logged in
            if "feed" in self.driver.current_url or "search" in self.driver.current_url:
                logger.info("Successfully logged in!")
                return True
            elif "checkpoint" in self.driver.current_url:
                logger.warning("Verification required. Please complete it manually.")
                if not self.headless:
                    input("Press Enter after completing verification...")
                return True
            
            return False
            
        except Exception as e:
            logger.error(f"Login failed: {e}")
            return False
    
    def search_profiles(self, search_url: str = None) -> List[str]:
        """Get profile URLs from search results"""
        if not search_url:
            search_url = "https://www.linkedin.com/search/results/people/?keywords=oracle%20plsql%20developer"
        
        logger.info(f"Navigating to search: {search_url}")
        self.driver.get(search_url)
        time.sleep(3)
        
        profile_urls = []
        
        try:
            # Wait for results
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "reusable-search__result-container"))
            )
            
            # Scroll to load more results
            for _ in range(3):
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(2)
            
            # Get all profile links
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            profile_containers = soup.find_all('li', class_='reusable-search__result-container')
            
            for container in profile_containers:
                link = container.find('a', class_='app-aware-link')
                if link and link.get('href'):
                    url = link['href']
                    if '/in/' in url:
                        if not url.startswith('http'):
                            url = f"https://www.linkedin.com{url}"
                        profile_urls.append(url.split('?')[0])  # Clean URL
            
            logger.info(f"Found {len(profile_urls)} profiles")
            
        except Exception as e:
            logger.error(f"Error getting search results: {e}")
        
        return profile_urls
    
    def extract_profile_data(self, profile_url: str) -> ProfileData:
        """Extract detailed data from a profile"""
        logger.info(f"Extracting data from: {profile_url}")
        self.driver.get(profile_url)
        time.sleep(3)
        
        profile = ProfileData(name="Unknown", profile_url=profile_url)
        
        try:
            # Wait for profile to load
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "text-heading-xlarge"))
            )
            
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            
            # Extract name
            name_elem = soup.find('h1', class_='text-heading-xlarge')
            if name_elem:
                profile.name = name_elem.get_text().strip()
            
            # Extract headline
            headline_elem = soup.find('div', class_='text-body-medium')
            if headline_elem:
                profile.headline = headline_elem.get_text().strip()
            
            # Extract location
            location_elem = soup.find('span', class_='text-body-small inline t-black--light break-words')
            if location_elem:
                profile.location = location_elem.get_text().strip()
            
            # Extract current company from experience
            exp_section = soup.find('div', id='experience')
            if exp_section:
                first_exp = exp_section.find('div', class_='display-flex align-items-center')
                if first_exp:
                    company_elem = first_exp.find('span', class_='t-14 t-normal')
                    if company_elem:
                        profile.current_company = company_elem.get_text().strip().split('·')[0]
                
                # Get experience summary
                exp_items = exp_section.find_all('li', class_='artdeco-list__item')[:3]
                experiences = []
                for exp in exp_items:
                    role = exp.find('div', class_='display-flex align-items-center')
                    if role:
                        role_text = role.get_text().strip().replace('\n', ' ')
                        experiences.append(role_text)
                profile.experience = ' | '.join(experiences)[:500]
            
            # Extract about section
            about_section = soup.find('div', class_='pv-shared-text-with-see-more')
            if about_section:
                profile.about = about_section.get_text().strip()
            
            # Extract education
            edu_section = soup.find('div', id='education')
            if edu_section:
                first_edu = edu_section.find('li', class_='artdeco-list__item')
                if first_edu:
                    edu_text = first_edu.get_text().strip().replace('\n', ' ')
                    profile.education = edu_text[:200]
            
            # Extract connections
            conn_elem = soup.find('span', class_='t-bold')
            if conn_elem and 'connection' in conn_elem.get_text().lower():
                profile.connections = conn_elem.get_text().strip()
            
            # Click to show all skills
            try:
                show_skills_btn = self.driver.find_element(By.XPATH, "//a[contains(@href, '/details/skills/')]")
                self.driver.execute_script("arguments[0].click();", show_skills_btn)
                time.sleep(2)
                
                # Re-parse for skills
                soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            except:
                logger.info("Could not expand skills section")
            
            # Extract skills
            skills = []
            skills_section = soup.find('div', {'data-view-name': 'profile-card'})
            if not skills_section:
                skills_section = soup.find('section', class_='artdeco-card')
            
            if skills_section:
                skill_elements = skills_section.find_all('div', class_='display-flex align-items-center')
                for elem in skill_elements:
                    skill_name = elem.find('span', {'aria-hidden': 'true'})
                    if skill_name:
                        skill_text = skill_name.get_text().strip()
                        if skill_text and len(skill_text) < 50:  # Filter out non-skill text
                            skills.append(skill_text)
            
            profile.skills = skills
            profile.total_skills_count = len(skills)
            
            # Calculate skill match score
            matched_skills = []
            if skills:
                for target_skill in self.target_skills:
                    for skill in skills:
                        if target_skill.lower() in skill.lower():
                            matched_skills.append(target_skill)
                            break
            
            profile.required_skills_matched = list(set(matched_skills))
            profile.skill_match_score = (len(profile.required_skills_matched) / len(self.target_skills)) * 100
            
            # Try to extract email (usually in contact info)
            try:
                contact_btn = self.driver.find_element(By.ID, "top-card-text-details-contact-info")
                contact_btn.click()
                time.sleep(2)
                
                contact_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
                email_section = contact_soup.find('section', class_='ci-email')
                if email_section:
                    email_link = email_section.find('a')
                    if email_link:
                        profile.email = email_link.get_text().strip()
                
                # Close modal
                close_btn = self.driver.find_element(By.XPATH, "//button[@aria-label='Dismiss']")
                close_btn.click()
            except:
                logger.info("Could not extract contact info")
            
        except Exception as e:
            logger.error(f"Error extracting profile data: {e}")
        
        return profile
    
    def scrape_profiles(self, search_url: str = None, max_profiles: int = 10) -> List[ProfileData]:
        """Main method to scrape profiles"""
        profiles_data = []
        
        try:
            # Setup and login
            self.setup_driver()
            
            if not self.login():
                raise Exception("Failed to login to LinkedIn")
            
            # Get profile URLs
            profile_urls = self.search_profiles(search_url)
            
            # Limit profiles
            profile_urls = profile_urls[:max_profiles]
            
            # Extract data from each profile
            for i, url in enumerate(profile_urls, 1):
                logger.info(f"Processing profile {i}/{len(profile_urls)}")
                
                try:
                    profile_data = self.extract_profile_data(url)
                    profiles_data.append(profile_data)
                    
                    # Delay between profiles
                    time.sleep(2)
                    
                except Exception as e:
                    logger.error(f"Error processing profile {url}: {e}")
                    continue
            
            # Sort by match score
            profiles_data.sort(key=lambda x: x.skill_match_score, reverse=True)
            
        except Exception as e:
            logger.error(f"Scraping failed: {e}")
        
        finally:
            if self.driver:
                self.driver.quit()
        
        return profiles_data
    
    def export_to_excel(self, profiles: List[ProfileData], filename: str = None):
        """Export profiles to Excel with formatting"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"linkedin_profiles_{timestamp}.xlsx"
        
        # Convert to DataFrame
        data = [profile.to_dict() for profile in profiles]
        df = pd.DataFrame(data)
        
        # Create Excel writer with formatting
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Profiles', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Profiles']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add color coding for match scores
            from openpyxl.styles import PatternFill
            from openpyxl.formatting.rule import CellIsRule
            
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
            red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            
            # Find the Match Score column
            for idx, col in enumerate(df.columns, 1):
                if 'Match Score' in col:
                    col_letter = worksheet.cell(row=1, column=idx).column_letter
                    
                    # Apply conditional formatting
                    worksheet.conditional_formatting.add(
                        f'{col_letter}2:{col_letter}{len(df)+1}',
                        CellIsRule(operator='greaterThan', formula=['70'], fill=green_fill)
                    )
                    worksheet.conditional_formatting.add(
                        f'{col_letter}2:{col_letter}{len(df)+1}',
                        CellIsRule(operator='between', formula=['40', '70'], fill=yellow_fill)
                    )
                    worksheet.conditional_formatting.add(
                        f'{col_letter}2:{col_letter}{len(df)+1}',
                        CellIsRule(operator='lessThan', formula=['40'], fill=red_fill)
                    )
        
        logger.info(f"Data exported to {filename}")
        return filename


def main():
    """Main execution function"""
    print("\n=== LinkedIn Profile Scraper ===\n")
    
    # Get credentials
    email = input("Enter LinkedIn email (or press Enter to use env variable): ").strip()
    if not email:
        email = os.getenv('LINKEDIN_EMAIL')
    
    password = input("Enter LinkedIn password (or press Enter to use env variable): ").strip()
    if not password:
        password = os.getenv('LINKEDIN_PASSWORD')
    
    if not email or not password:
        print("❌ LinkedIn credentials are required!")
        return
    
    # Get search parameters
    search_url = input("\nEnter LinkedIn search URL (or press Enter for default Oracle PL/SQL search): ").strip()
    if not search_url:
        search_url = "https://www.linkedin.com/search/results/people/?keywords=oracle%20plsql%20developer"
    
    max_profiles = input("How many profiles to scrape? (default: 10): ").strip()
    max_profiles = int(max_profiles) if max_profiles else 10
    
    # Run scraper
    scraper = LinkedInProfileScraper(email=email, password=password, headless=False)
    
    print(f"\n🔍 Starting to scrape {max_profiles} profiles...")
    profiles = scraper.scrape_profiles(search_url, max_profiles)
    
    if profiles:
        print(f"\n✅ Successfully scraped {len(profiles)} profiles!")
        
        # Export to Excel
        filename = scraper.export_to_excel(profiles)
        print(f"📊 Data exported to: {filename}")
        
        # Show top matches
        print("\n🏆 Top 5 Matches:")
        for i, profile in enumerate(profiles[:5], 1):
            print(f"\n{i}. {profile.name}")
            print(f"   Score: {profile.skill_match_score:.1f}%")
            print(f"   Skills: {', '.join(profile.required_skills_matched[:5]) if profile.required_skills_matched else 'None'}")
            print(f"   Company: {profile.current_company}")
    else:
        print("❌ No profiles were scraped")


if __name__ == "__main__":
    main()